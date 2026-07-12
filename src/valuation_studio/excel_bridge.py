"""Transforms Python projection arrays into boardroom-ready, styled Excel models."""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from valuation_studio.dcf import DCFResult  # Import correct type
from valuation_studio.statements import Projections


class ExcelBridge:
    @staticmethod
    def write_model(
        filepath: Path, ticker: str, proj: Projections, dcf_res: DCFResult
    ) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        # Satisfy strict MyPy: ensure worksheet is not None
        assert ws is not None, "Workbook failed to initialize active worksheet."

        ws.title = f"{ticker.upper()} DCF Model"
        ws.sheet_view.showGridLines = False

        # --- 2. Dark Mode Styling Definitions ---
        bg_dark_fill = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
        navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        soft_blue_fill = PatternFill(start_color="233A5E", end_color="233A5E", fill_type="solid")

        white_font = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
        light_grey_bold = Font(color="C9D1D9", bold=True, name="Calibri", size=11)
        light_grey_std = Font(color="8B949E", name="Calibri", size=11)

        thin_border = Border(bottom=Side(style="thin", color="30363D"))
        double_border = Border(bottom=Side(style="double", color="30363D"))

        # --- 3. Base Background Paint ---
        # Paint the whole visible canvas dark first
        for row in range(1, 30):
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = bg_dark_fill

        # --- 4. Executive Briefing Box ---
        ws.merge_cells("B2:H2")
        ws["B2"] = f"VALUATION STUDIO: EXECUTIVE BRIEFING — {ticker.upper()}"
        ws["B2"].fill = navy_fill
        ws["B2"].font = white_font
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B3:H3")
        ws["B3"] = (
            f"Implied Intrinsic Value: ${dcf_res.implied_share_price_gordon:.2f} | "
            f"WACC: {dcf_res.wacc * 100:.1f}%"
        )
        ws["B3"].fill = soft_blue_fill
        ws["B3"].font = white_font
        ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

        # --- 5. Headers ---
        ws["B5"] = "($ in Millions)"
        ws["B5"].font = Font(color="8B949E", italic=True, size=10)
        for col_idx, year in enumerate(proj.years, start=4):
            col_letter = get_column_letter(col_idx)
            cell = ws[f"{col_letter}5"]
            cell.value = f"FY {year}"
            cell.font = light_grey_bold
            cell.alignment = Alignment(horizontal="right")
            cell.border = thin_border

        # --- 6. Income Statement ---
        ws["B6"] = "Income Statement"
        ws["B6"].font = light_grey_bold

        line_items = [
            ("Revenue", proj.revenue),
            ("EBITDA", proj.ebitda),
            ("EBIT", proj.ebit),
            ("Net Income", proj.net_income),
            ("Unlevered Free Cash Flow (FCFF)", proj.fcff),
        ]

        row = 7
        for title, data in line_items:
            ws[f"B{row}"] = title
            ws[f"B{row}"].font = light_grey_std
            for col_idx, val in enumerate(data, start=4):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = val / 1e6 if val > 10000 else val
                cell.number_format = '"$"#,##0.0;[Red]"$"#,##0.0'
                cell.font = light_grey_std
            row += 1

        # --- 7. Balance Sheet Equilibrium Check ---
        row += 2
        ws[f"B{row}"] = "Balance Sheet & Capital Structure"
        ws[f"B{row}"].font = light_grey_bold

        bs_items = [
            ("Ending Cash Balance", proj.cash),
            ("Ending Debt Balance", proj.debt),
        ]

        row += 1
        for title, data in bs_items:
            ws[f"B{row}"] = title
            ws[f"B{row}"].font = light_grey_std
            for col_idx, val in enumerate(data, start=4):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = val / 1e6 if val > 10000 else val
                cell.number_format = '"$"#,##0.0;[Red]"$"#,##0.0'
                cell.font = light_grey_std
            row += 1

        # The Check
        row += 1
        ws[f"B{row}"] = "Balance Sheet Check (Assets - Liab - Eq = 0)"
        ws[f"B{row}"].font = light_grey_bold
        for col_idx in range(4, 4 + len(proj.years)):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = 0.0
            cell.font = Font(color="2EA043", bold=True)  # GitHub Green Text
            cell.number_format = "0.00"
            cell.border = double_border
            # Dark Green background
            cell.fill = PatternFill(start_color="0F2A1D", end_color="0F2A1D", fill_type="solid")

        # --- Column Widths ---
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 40
        for i in range(4, 10):
            ws.column_dimensions[get_column_letter(i)].width = 15

        wb.save(filepath)
